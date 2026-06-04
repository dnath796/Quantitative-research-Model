"""
WTI Crude Oil Return Drivers — Full Pipeline
MFIN8852 Financial Econometrics Final Project

Pipeline:  EDA  →  ADF stationarity  →  OLS  →  Diagnostics  →  HAC  →  Excel output
Sample: December 1990 – December 2025
Dependent variable: wti_chg_pct (monthly % change in WTI price)

Outputs:
  residual_diagnostics.png
  coefficient_plot.png
  wti_regression_results.xlsx   ← full workbook with OLS and HAC side-by-side
"""

import datetime
import warnings
warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import statsmodels.api as sm
from statsmodels.tsa.stattools import adfuller, kpss
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.stats.diagnostic import (het_breuschpagan,
                                          acorr_breusch_godfrey,
                                          linear_reset)
from statsmodels.stats.stattools import durbin_watson
from scipy import stats
from scipy.stats import f as f_dist


# ═══════════════════════════════════════════════════════════════════════════════
# 0.  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH   = 'master_datasheet.xlsx'
SHEET_NAME   = 'Combined Monthly'
SAMPLE_START = '2015-12'
SAMPLE_END   = '2025-12'
OUTPUT_XLSX  = 'wti_regression_results.xlsx'

TARGET_COLS = [
    'wti_chg_pct',
    'dxy_chg_pct',
    'sp500_chg_pct',
    'gld_chg_pct',
    'copper_chg_pct',
    'ng_chg_pct',
    'us_crudeProd_chg_pct',
    'indpro_chg_pct',
    'crude_stk_excl_spr',
    'spr_inv_chg',
    'tcmsy10_rate_1st_diff',
    'vix_close',
    'refin_util_rate',
    'igrea_level',
    'gpr_idx_level',
]

LABELS = {
    'wti_chg_pct':           'WTI Return (%)',
    'dxy_chg_pct':           'DXY Return (%)',
    'sp500_chg_pct':         'S&P 500 Return (%)',
    'gld_chg_pct':           'Gold Return (%)',
    'copper_chg_pct':        'Copper Return (%)',
    'ng_chg_pct':            'Natural Gas Return (%)',
    'us_crudeProd_chg_pct':  'US Crude Prod. Chg (%)',
    'indpro_chg_pct':        'INDPRO Chg (%)',
    'crude_stk_excl_spr':    'Commercial Crude Stocks (Level)',
    'spr_inv_chg':           'SPR Inventory Chg',
    'tcmsy10_rate_1st_diff': '10-Yr Treasury Delta (pp)',
    'vix_close':             'VIX (Level)',
    'refin_util_rate':       'Refinery Utilization Rate',
    'igrea_level':           'Kilian IGREA (Level)',
    'gpr_idx_level':         'GPR Index (Level)',
}

# Regressors EXCLUDING refinery utilization (VIF=13) and target
REGRESSORS = [c for c in TARGET_COLS
              if c not in ('wti_chg_pct', 'refin_util_rate')]

# ── Style helpers ─────────────────────────────────────────────────────────────

def _hdr(bold=True, color='FFFFFF', bg='2F5496', size=10, italic=False):
    return {'font': Font(bold=bold, color=color, size=size, italic=italic,
                         name='Arial'),
            'fill': PatternFill('solid', start_color=bg),
            'align': Alignment(horizontal='center', vertical='center',
                               wrap_text=True)}

def _cell(bold=False, color='000000', bg=None, align='center', size=10,
          italic=False, wrap=False):
    fill = PatternFill('solid', start_color=bg) if bg else PatternFill()
    return {'font': Font(bold=bold, color=color, size=size, italic=italic,
                         name='Arial'),
            'fill': fill,
            'align': Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap)}

THIN = Border(
    left=Side(style='thin'),  right=Side(style='thin'),
    top=Side(style='thin'),   bottom=Side(style='thin'),
)
MEDIUM_BOTTOM = Border(bottom=Side(style='medium'))

def apply(ws, row, col, value, style: dict, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if 'font'  in style: c.font      = style['font']
    if 'fill'  in style: c.fill      = style['fill']
    if 'align' in style: c.alignment = style['align']
    c.border = THIN
    if num_fmt: c.number_format = num_fmt
    return c

def merge_hdr(ws, r1, c1, r2, c2, value, style):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if 'font'  in style: c.font      = style['font']
    if 'fill'  in style: c.fill      = style['fill']
    if 'align' in style: c.alignment = style['align']
    c.border = THIN

SIG_STARS = lambda p: '***' if p < 0.01 else ('**' if p < 0.05
                         else ('*' if p < 0.1 else ''))


# ═══════════════════════════════════════════════════════════════════════════════
# 1.  LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════

wb_in = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws_in = wb_in[SHEET_NAME]
rows  = list(ws_in.iter_rows(values_only=True))
wb_in.close()

field_row = rows[2]
col_idx   = {f: i for i, f in enumerate(field_row) if f is not None}

records = []
for r in rows[3:]:
    if not isinstance(r[0], datetime.datetime):
        continue
    rec = {'Date': r[0]}
    for col in TARGET_COLS:
        rec[col] = r[col_idx[col]]
    records.append(rec)

df = pd.DataFrame(records).set_index('Date')
df.index = pd.to_datetime(df.index)
df = df.apply(pd.to_numeric, errors='coerce').dropna()
df = df.loc[SAMPLE_START:SAMPLE_END]
T  = len(df)

print(f'Sample: {df.index[0]:%Y-%m} to {df.index[-1]:%Y-%m}  |  T = {T}')

y     = df['wti_chg_pct']
X_raw = df[REGRESSORS]
X     = sm.add_constant(X_raw)


# ═══════════════════════════════════════════════════════════════════════════════
# 2.  EDA — DESCRIPTIVE STATISTICS  (printed; also goes to Excel)
# ═══════════════════════════════════════════════════════════════════════════════

desc = df[TARGET_COLS].describe().T
desc['skew']     = df[TARGET_COLS].skew()
desc['kurtosis'] = df[TARGET_COLS].kurtosis()
print('\n── Descriptive Statistics ──')
print(desc[['mean', 'std', 'min', 'max', 'skew', 'kurtosis']].to_string())


# ═══════════════════════════════════════════════════════════════════════════════
# 3.  STATIONARITY — ADF + KPSS
# ═══════════════════════════════════════════════════════════════════════════════

def run_adf(series, name):
    r = adfuller(series.dropna(), autolag='AIC', regression='c')
    return {'variable': name, 'adf_stat': r[0], 'adf_p': r[1],
            'adf_lags': r[2], 'adf_sig': SIG_STARS(r[1]),
            'adf_conclusion': 'Stationary' if r[1] < 0.05 else 'Unit Root'}

def run_kpss(series, name):
    stat, p, lags, crit = kpss(series.dropna(), regression='c',
                                nlags='auto')
    # KPSS: H0 = stationary; reject H0 means non-stationary
    conclusion = 'Non-Stationary' if p < 0.05 else 'Stationary'
    return {'kpss_stat': stat, 'kpss_p': p, 'kpss_lags': lags,
            'kpss_sig': SIG_STARS(p), 'kpss_conclusion': conclusion}

stat_rows = []
all_stat_cols = TARGET_COLS  # include refin_util_rate in ADF table
for col in all_stat_cols:
    adf  = run_adf(df[col], LABELS[col])
    kpss_res = run_kpss(df[col], LABELS[col])
    both_agree = (adf['adf_conclusion'] == 'Stationary' and
                  kpss_res['kpss_conclusion'] == 'Stationary')
    stat_rows.append({**adf, **kpss_res,
                      'overall': 'Stationary ✓' if both_agree
                                 else 'Conflict / Review'})

stat_df = pd.DataFrame(stat_rows)
print('\n── Stationarity Summary ──')
print(stat_df[['variable','adf_stat','adf_p','adf_conclusion',
               'kpss_stat','kpss_p','kpss_conclusion','overall']].to_string())


# ═══════════════════════════════════════════════════════════════════════════════
# 4.  OLS ESTIMATION  (plain OLS — BEFORE correction)
# ═══════════════════════════════════════════════════════════════════════════════

ols = sm.OLS(y, X).fit()
resid  = ols.resid
fitted = ols.fittedvalues

print('\n' + '='*72)
print('OLS RESULTS (before HAC correction)')
print('='*72)
print(f'  R²={ols.rsquared:.4f}  Adj-R²={ols.rsquared_adj:.4f}  '
      f'F={ols.fvalue:.3f}  p={ols.f_pvalue:.4f}')


# ═══════════════════════════════════════════════════════════════════════════════
# 5.  DIAGNOSTIC TESTS
# ═══════════════════════════════════════════════════════════════════════════════

# 5a. VIF
vif_vals = {}
for i, col in enumerate(X_raw.columns):
    vif_vals[col] = variance_inflation_factor(X_raw.values, i)

# 5b. Breusch-Pagan
bp_lm, bp_pval, bp_f, bp_fpval = het_breuschpagan(resid, X)

# 5c. Breusch-Godfrey
bg_lm, bg_pval, bg_f, bg_fpval = acorr_breusch_godfrey(ols, nlags=12)

# 5d. Durbin-Watson
dw = durbin_watson(resid)

# 5e. Jarque-Bera
jb_stat, jb_pval = stats.jarque_bera(resid)
skew_val  = float(stats.skew(resid))
kurt_val  = float(stats.kurtosis(resid))

# 5f. RESET test
reset_result = linear_reset(ols, power=3, use_f=True)
reset_f = reset_result.fvalue
reset_p = reset_result.pvalue

# 5g. Chow test — structural break 2014-01 (shale revolution)
break_date = '2014-01'
df_pre  = df.loc[:break_date]
df_post = df.loc[break_date:]
ols_pre  = sm.OLS(df_pre['wti_chg_pct'],
                  sm.add_constant(df_pre[REGRESSORS])).fit()
ols_post = sm.OLS(df_post['wti_chg_pct'],
                  sm.add_constant(df_post[REGRESSORS])).fit()
k = len(REGRESSORS) + 1
chow_num = (ols.ssr - ols_pre.ssr - ols_post.ssr) / k
chow_den = (ols_pre.ssr + ols_post.ssr) / (T - 2 * k)
chow_f   = chow_num / chow_den
chow_p   = 1 - f_dist.cdf(chow_f, k, T - 2 * k)

print(f'\n  Breusch-Pagan   : LM={bp_lm:.4f}  p={bp_pval:.4f}')
print(f'  Breusch-Godfrey : LM={bg_lm:.4f}  p={bg_pval:.4f}')
print(f'  Durbin-Watson   : {dw:.4f}')
print(f'  Jarque-Bera     : JB={jb_stat:.4f}  p={jb_pval:.4f}')
print(f'  RESET           : F={reset_f:.4f}  p={reset_p:.4f}')
print(f'  Chow (2014-01)  : F={chow_f:.4f}  p={chow_p:.4f}')


# ═══════════════════════════════════════════════════════════════════════════════
# 6.  HAC CORRECTION  (Newey-West, AFTER correction)
# ═══════════════════════════════════════════════════════════════════════════════

needs_hac = bp_pval < 0.05 or bg_pval < 0.05
nw_lags   = int(np.ceil(4 * (T / 100) ** (2 / 9)))

if needs_hac:
    ols_hac    = sm.OLS(y, X).fit(cov_type='HAC',
                                  cov_kwds={'maxlags': nw_lags})
    final_model = ols_hac
    print(f'\nNewey-West HAC applied (lags={nw_lags})')
else:
    ols_hac     = ols        # alias so later code still runs
    final_model = ols
    print('\nNo HAC needed — OLS standard errors are valid.')


# ═══════════════════════════════════════════════════════════════════════════════
# 7.  STANDARDISED (BETA) COEFFICIENTS
# ═══════════════════════════════════════════════════════════════════════════════

X_std   = (X_raw - X_raw.mean()) / X_raw.std()
y_std   = (y     - y.mean())     / y.std()
ols_std = sm.OLS(y_std, sm.add_constant(X_std)).fit()
betas   = ols_std.params.drop('const')


# ═══════════════════════════════════════════════════════════════════════════════
# 8.  RESIDUAL DIAGNOSTIC PLOTS
# ═══════════════════════════════════════════════════════════════════════════════

fig = plt.figure(figsize=(14, 10))
gs  = gridspec.GridSpec(2, 3, figure=fig, hspace=0.4, wspace=0.35)

ax1 = fig.add_subplot(gs[0, :2])
ax1.plot(df.index, resid, linewidth=0.8, color='steelblue')
ax1.axhline(0, color='red', linewidth=0.8, linestyle='--')
ax1.set_title('OLS Residuals Over Time', fontweight='bold')
ax1.set_xlabel('Date'); ax1.set_ylabel('Residual')
ax1.grid(True, alpha=0.3)

ax2 = fig.add_subplot(gs[0, 2])
ax2.scatter(fitted, resid, alpha=0.4, s=12, color='steelblue')
ax2.axhline(0, color='red', linewidth=0.8, linestyle='--')
ax2.set_title('Residuals vs Fitted', fontweight='bold')
ax2.set_xlabel('Fitted Values'); ax2.set_ylabel('Residual')
ax2.grid(True, alpha=0.3)

ax3 = fig.add_subplot(gs[1, 0])
ax3.hist(resid, bins=35, color='steelblue', edgecolor='white', alpha=0.8)
xr = np.linspace(resid.min(), resid.max(), 200)
ax3.plot(xr,
         stats.norm.pdf(xr, resid.mean(), resid.std()) *
         len(resid) * (resid.max() - resid.min()) / 35,
         color='red', linewidth=1.5, label='Normal')
ax3.set_title('Residual Distribution', fontweight='bold')
ax3.set_xlabel('Residual'); ax3.legend(fontsize=8)
ax3.grid(True, alpha=0.3)

ax4 = fig.add_subplot(gs[1, 1])
(osm, osr), (slope, intercept, _) = stats.probplot(resid, dist='norm')
ax4.scatter(osm, osr, alpha=0.4, s=12, color='steelblue')
ax4.plot(osm, slope * np.array(osm) + intercept, color='red', linewidth=1.5)
ax4.set_title('Q-Q Plot of Residuals', fontweight='bold')
ax4.set_xlabel('Theoretical Quantiles'); ax4.set_ylabel('Sample Quantiles')
ax4.grid(True, alpha=0.3)

ax5 = fig.add_subplot(gs[1, 2])
ax5.scatter(fitted, y, alpha=0.4, s=12, color='steelblue')
ax5.plot([y.min(), y.max()], [y.min(), y.max()],
         color='red', linewidth=1.5, linestyle='--')
ax5.set_title('Actual vs Fitted WTI Return', fontweight='bold')
ax5.set_xlabel('Fitted'); ax5.set_ylabel('Actual')
ax5.grid(True, alpha=0.3)

fig.suptitle('OLS Residual Diagnostics', fontsize=13, fontweight='bold')
plt.savefig('residual_diagnostics.png', dpi=150, bbox_inches='tight')
plt.close('all')
print('\nSaved: residual_diagnostics.png')


# ═══════════════════════════════════════════════════════════════════════════════
# 9.  COEFFICIENT PLOT  (HAC confidence intervals)
# ═══════════════════════════════════════════════════════════════════════════════

params  = final_model.params.drop('const')
ci_low  = final_model.conf_int().loc[params.index, 0]
ci_high = final_model.conf_int().loc[params.index, 1]
pvals   = final_model.pvalues.drop('const')
short_labels = [LABELS[v].replace(' (%)', '').replace(' (Level)', '')
                          .replace(' (pp)', ' Δ') for v in params.index]
colors = ['#d62728' if p < 0.05 else '#aec7e8' for p in pvals]

fig, ax = plt.subplots(figsize=(10, 7))
y_pos   = range(len(params))
ax.barh(y_pos, params.values,
        xerr=[params.values - ci_low.values,
              ci_high.values - params.values],
        color=colors, edgecolor='black', linewidth=0.5,
        capsize=4, error_kw={'linewidth': 1})
ax.set_yticks(list(y_pos)); ax.set_yticklabels(short_labels, fontsize=9)
ax.axvline(0, color='black', linewidth=0.8)
ax.set_xlabel('Coefficient (effect on monthly WTI return)')
ax.set_title('OLS Coefficients with 95% Confidence Intervals\n'
             '(red = significant at 5%, blue = insignificant)',
             fontweight='bold')
ax.grid(True, axis='x', alpha=0.3)
plt.tight_layout()
plt.savefig('coefficient_plot.png', dpi=150, bbox_inches='tight')
plt.close('all')
print('Saved: coefficient_plot.png')


# ═══════════════════════════════════════════════════════════════════════════════
# 10.  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
#
# Sheets:
#   1. Summary          — model metadata, fit stats, diagnostic results
#   2. OLS vs HAC       — coefficients side-by-side (before & after)
#   3. Stationarity     — ADF + KPSS table
#   4. VIF              — variance inflation factors
#   5. Diagnostics      — all test statistics in one place
#   6. Beta Coefs       — standardised coefficients + rank
#   7. Descriptive Stats— mean/std/min/max/skew/kurt for all series
#   8. Raw Data         — the cleaned monthly panel used in estimation
# ═══════════════════════════════════════════════════════════════════════════════

wb_out = Workbook()

# colour palette
BLUE_DARK   = '2F5496'
BLUE_MID    = 'D9E1F2'
BLUE_LIGHT  = 'EEF2F9'
GREEN_DARK  = '375623'
GREEN_LIGHT = 'E2EFDA'
RED_DARK    = 'C00000'
RED_LIGHT   = 'FFDCE1'
ORANGE      = 'F4B942'
GREY_HDR    = '595959'
WHITE       = 'FFFFFF'
YELLOW_HIGH = 'FFEB84'


# ─── helper: auto-fit columns ────────────────────────────────────────────────
def autofit(ws, min_w=8, max_w=40):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb_out.active
ws1.title = 'Summary'
ws1.sheet_view.showGridLines = False

# Title block
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 18
merge_hdr(ws1, 1, 1, 1, 6,
          'WTI Crude Oil Return Drivers — Regression Summary',
          {**_hdr(size=13), 'fill': PatternFill('solid', start_color=BLUE_DARK)})
merge_hdr(ws1, 2, 1, 2, 6,
          'MFIN8852 Financial Econometrics  |  Dec 1990 – Dec 2025  |  T = 421',
          {**_hdr(size=10, bold=False),
           'fill': PatternFill('solid', start_color=GREY_HDR)})

# ── Model Fit ────────────────────────────────────────────────────────────────
r = 4
merge_hdr(ws1, r, 1, r, 6, 'MODEL FIT STATISTICS',
          _hdr(bg=BLUE_DARK))
r += 1

fit_rows = [
    ('Observations (T)',        T,                   None),
    ('Regressors (excl. const)',len(REGRESSORS),     None),
    ('R-squared',               ols.rsquared,        '0.0000'),
    ('Adjusted R-squared',      ols.rsquared_adj,    '0.0000'),
    ('F-statistic',             ols.fvalue,          '0.000'),
    ('F p-value',               ols.f_pvalue,        '0.0000'),
    ('AIC',                     ols.aic,             '0.00'),
    ('BIC',                     ols.bic,             '0.00'),
    ('Log-Likelihood',          ols.llf,             '0.00'),
    ('HAC lags (Newey-West)',   nw_lags if needs_hac else 'N/A', None),
]

even = PatternFill('solid', start_color=BLUE_LIGHT)
odd  = PatternFill()
for i, (label, val, fmt) in enumerate(fit_rows):
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    apply(ws1, r, 1, label, _cell(bold=True, align='left', bg=bg))
    ws1.merge_cells(start_row=r, start_column=1,
                    end_row=r,   end_column=3)
    c = ws1.cell(row=r, column=4, value=val)
    c.font      = Font(name='Arial', size=10)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center')
    c.border    = THIN
    if fmt and isinstance(val, (int, float)): c.number_format = fmt
    ws1.merge_cells(start_row=r, start_column=4,
                    end_row=r,   end_column=6)
    r += 1

# ── Diagnostic Tests ─────────────────────────────────────────────────────────
r += 1
merge_hdr(ws1, r, 1, r, 6, 'DIAGNOSTIC TESTS', _hdr(bg=BLUE_DARK))
r += 1

# sub-header
for col, txt in enumerate(
        ['Test', 'Statistic', 'p-value', 'Decision', 'Action', 'Notes'], 1):
    apply(ws1, r, col, txt, _hdr(bg=GREY_HDR, size=9))
r += 1

diag_data = [
    ('Breusch-Pagan (heteroskedasticity)',
     f'LM = {bp_lm:.4f}', f'{bp_pval:.4f}',
     'Reject H0' if bp_pval < 0.05 else 'Fail to Reject',
     'HAC applied' if bp_pval < 0.05 else 'None',
     'Residual variance non-constant'),
    ('Breusch-Godfrey (serial correlation, 12 lags)',
     f'LM = {bg_lm:.4f}', f'{bg_pval:.4f}',
     'Reject H0' if bg_pval < 0.05 else 'Fail to Reject',
     'HAC applied' if bg_pval < 0.05 else 'None',
     'Higher-order autocorrelation'),
    ('Durbin-Watson (1st-order autocorrel.)',
     f'{dw:.4f}', '—',
     'No 1st-order autocorr.' if 1.5 < dw < 2.5 else 'Review',
     'None', 'Complement to B-G'),
    ('Jarque-Bera (normality)',
     f'JB = {jb_stat:.2f}', f'{jb_pval:.4f}',
     'Reject H0' if jb_pval < 0.05 else 'Fail to Reject',
     'CLT (T=421 sufficient)',
     f'Skew={skew_val:.3f}, Kurt={kurt_val:.3f}'),
    ('Ramsey RESET (functional form)',
     f'F = {reset_f:.4f}', f'{reset_p:.4f}',
     'Reject H0' if reset_p < 0.05 else 'Fail to Reject',
     'Review spec.' if reset_p < 0.05 else 'None',
     'Linear spec. adequate' if reset_p >= 0.05 else 'Possible non-linearity'),
    (f'Chow Test (break @ {break_date})',
     f'F = {chow_f:.4f}', f'{chow_p:.4f}',
     'Reject H0' if chow_p < 0.05 else 'Fail to Reject',
     'Structural break detected' if chow_p < 0.05 else 'Stable coefficients',
     'Pre/Post shale revolution'),
]

for i, row_data in enumerate(diag_data):
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    is_reject = 'Reject' in row_data[3]
    for col, val in enumerate(row_data, 1):
        st = _cell(align='left' if col in (1, 4, 5, 6) else 'center',
                   bg=bg)
        if col == 4:  # Decision — colour by outcome
            st = _cell(bold=True, align='center',
                       color=RED_DARK if is_reject else GREEN_DARK,
                       bg=bg)
        apply(ws1, r, col, val, st)
    r += 1

# ── VIF mini-table ────────────────────────────────────────────────────────────
r += 1
merge_hdr(ws1, r, 1, r, 3, 'VARIANCE INFLATION FACTORS', _hdr(bg=BLUE_DARK))
merge_hdr(ws1, r, 4, r, 6, 'Max VIF after dropping refin_util_rate: '
          f'{max(vif_vals.values()):.2f}',
          _hdr(bg=GREY_HDR, bold=False))
r += 1
for col, h in enumerate(['Variable', 'VIF', 'Status'], 1):
    apply(ws1, r, col, h, _hdr(bg=GREY_HDR, size=9))
r += 1
for i, (var, vif) in enumerate(sorted(vif_vals.items(),
                                       key=lambda x: -x[1])):
    bg = RED_LIGHT if vif > 10 else (YELLOW_HIGH if vif > 5 else WHITE)
    apply(ws1, r, 1, LABELS[var], _cell(align='left', bg=bg))
    apply(ws1, r, 2, round(vif, 4), _cell(bg=bg), '0.00')
    status = '⚠ HIGH' if vif > 10 else ('Moderate' if vif > 5 else 'OK ✓')
    apply(ws1, r, 3, status, _cell(bold=vif > 10,
          color=RED_DARK if vif > 10 else '000000', bg=bg))
    r += 1

for col in range(1, 7):
    ws1.column_dimensions[get_column_letter(col)].width = 26
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — OLS vs HAC (BEFORE & AFTER side-by-side)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet('OLS vs HAC')
ws2.sheet_view.showGridLines = False

ws2.row_dimensions[1].height = 28
ws2.row_dimensions[2].height = 16
merge_hdr(ws2, 1, 1, 1, 12,
          'Regression Coefficients — OLS (Before) vs Newey-West HAC (After)',
          _hdr(size=12, bg=BLUE_DARK))
merge_hdr(ws2, 2, 1, 2, 12,
          f'Dependent variable: Monthly WTI Return (%)  |  T={T}  |  '
          f'HAC lags={nw_lags}  |  *** p<0.01  ** p<0.05  * p<0.10',
          _hdr(size=9, bold=False, bg=GREY_HDR))

# Column headers — row 3: section banners
merge_hdr(ws2, 3, 1,  3, 1,  'Variable',    _hdr(bg=BLUE_DARK))
merge_hdr(ws2, 3, 2,  3, 6,  'OLS — Standard Errors (Before Correction)',
          _hdr(bg='1F4E79'))
merge_hdr(ws2, 3, 7,  3, 11, f'Newey-West HAC (After Correction, lags={nw_lags})',
          _hdr(bg='375623'))
merge_hdr(ws2, 3, 12, 3, 12, 'Sig. Change?', _hdr(bg=GREY_HDR))

# Row 4: column sub-headers
sub_hdrs = ['Coefficient', 'Std. Error', 't-stat', 'p-value', 'Sig.']
apply(ws2, 4, 1, 'Variable', _hdr(bg=GREY_HDR, size=9))
for i, h in enumerate(sub_hdrs):
    apply(ws2, 4, 2 + i, h,  _hdr(bg='1F4E79', size=9))
    apply(ws2, 4, 7 + i, h,  _hdr(bg='375623', size=9))
apply(ws2, 4, 12, '★', _hdr(bg=GREY_HDR, size=9))

row = 5
for i, var in enumerate(['const'] + REGRESSORS):
    label = 'Constant' if var == 'const' else LABELS[var]
    bg    = BLUE_LIGHT if i % 2 == 0 else WHITE

    # OLS plain
    coef_ols = ols.params[var]
    se_ols   = ols.bse[var]
    t_ols    = ols.tvalues[var]
    p_ols    = ols.pvalues[var]
    sig_ols  = SIG_STARS(p_ols)

    # HAC
    coef_hac = ols_hac.params[var]
    se_hac   = ols_hac.bse[var]
    t_hac    = ols_hac.tvalues[var]
    p_hac    = ols_hac.pvalues[var]
    sig_hac  = SIG_STARS(p_hac)

    changed  = sig_ols != sig_hac
    sig_col  = RED_DARK if p_hac < 0.05 else GREY_HDR

    apply(ws2, row, 1, label, _cell(align='left', bg=bg, bold=(var=='const')))

    # OLS columns
    apply(ws2, row, 2,  coef_ols, _cell(bg=bg), '0.0000')
    apply(ws2, row, 3,  se_ols,   _cell(bg=bg), '0.0000')
    apply(ws2, row, 4,  t_ols,    _cell(bg=bg), '0.000')
    apply(ws2, row, 5,  p_ols,    _cell(bg=bg), '0.0000')
    apply(ws2, row, 6,  sig_ols,  _cell(bold=bool(sig_ols), bg=bg,
                                         color=RED_DARK if p_ols < 0.05
                                               else '000000'))

    # HAC columns — highlight significant rows green
    hac_bg = GREEN_LIGHT if p_hac < 0.05 else bg
    apply(ws2, row, 7,  coef_hac, _cell(bg=hac_bg), '0.0000')
    apply(ws2, row, 8,  se_hac,   _cell(bg=hac_bg), '0.0000')
    apply(ws2, row, 9,  t_hac,    _cell(bg=hac_bg), '0.000')
    apply(ws2, row, 10, p_hac,    _cell(bg=hac_bg), '0.0000')
    apply(ws2, row, 11, sig_hac,  _cell(bold=bool(sig_hac),
                                         color=RED_DARK if p_hac < 0.05
                                               else '000000', bg=hac_bg))
    change_txt = '← Changed' if changed else ''
    apply(ws2, row, 12, change_txt,
          _cell(bold=changed, color=RED_DARK if changed else '000000',
                bg=YELLOW_HIGH if changed else bg))
    row += 1

# ── Model fit footer ─────────────────────────────────────────────────────────
row += 1
merge_hdr(ws2, row, 1, row, 12, 'MODEL FIT', _hdr(bg=BLUE_DARK))
row += 1
fit_footer = [
    ('R-squared',       f'{ols.rsquared:.4f}',     f'{ols_hac.rsquared:.4f}'),
    ('Adj. R-squared',  f'{ols.rsquared_adj:.4f}', f'{ols_hac.rsquared_adj:.4f}'),
    ('F-statistic',     f'{ols.fvalue:.3f}',        f'{ols_hac.fvalue:.3f}'),
    ('F p-value',       f'{ols.f_pvalue:.4f}',      f'{ols_hac.f_pvalue:.4f}'),
    ('AIC',             f'{ols.aic:.2f}',            '—'),
    ('BIC',             f'{ols.bic:.2f}',            '—'),
    ('Note', 'Coefficients identical; only SEs change under HAC', ''),
]
for j, (lbl, v_ols, v_hac) in enumerate(fit_footer):
    bg = BLUE_LIGHT if j % 2 == 0 else WHITE
    is_note = lbl == 'Note'
    apply(ws2, row, 1,  lbl,   _cell(bold=True, align='left', bg=bg))
    merge_hdr(ws2, row, 2, row, 6,  v_ols,
              _cell(align='center', bg=bg, italic=is_note))
    merge_hdr(ws2, row, 7, row, 12, v_hac,
              _cell(align='center', bg=bg, italic=is_note))
    row += 1

col_widths = [30, 12, 12, 10, 10, 6, 12, 12, 10, 10, 6, 14]
for ci, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = 'A5'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — STATIONARITY (ADF + KPSS)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb_out.create_sheet('Stationarity')
ws3.sheet_view.showGridLines = False

merge_hdr(ws3, 1, 1, 1, 10,
          'Stationarity Tests — ADF + KPSS (Confirmatory)',
          _hdr(bg=BLUE_DARK, size=12))
merge_hdr(ws3, 2, 1, 2, 10,
          'ADF H0: unit root  |  KPSS H0: stationary  |  Both must agree for confirmation',
          _hdr(bg=GREY_HDR, bold=False, size=9))

merge_hdr(ws3, 3, 1, 3, 1, 'Variable',    _hdr(bg=GREY_HDR))
merge_hdr(ws3, 3, 2, 3, 5, 'Augmented Dickey-Fuller', _hdr(bg='1F4E79'))
merge_hdr(ws3, 3, 6, 3, 9, 'KPSS',                   _hdr(bg='375623'))
merge_hdr(ws3, 3, 10, 3, 10,'Overall',               _hdr(bg=BLUE_DARK))

for c, h in enumerate(['Stat.','p-value','Lags','Decision'], 2):
    apply(ws3, 4, c,     h, _hdr(bg='1F4E79', size=9))
for c, h in enumerate(['Stat.','p-value','Lags','Decision'], 6):
    apply(ws3, 4, c,     h, _hdr(bg='375623', size=9))
apply(ws3, 4, 1,  'Variable', _hdr(bg=GREY_HDR, size=9))
apply(ws3, 4, 10, 'Result',   _hdr(bg=BLUE_DARK, size=9))

for i, sr in enumerate(stat_rows):
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    r  = 5 + i
    adf_ok  = sr['adf_conclusion'] == 'Stationary'
    kpss_ok = sr['kpss_conclusion'] == 'Stationary'
    both_ok = adf_ok and kpss_ok

    apply(ws3, r, 1,  sr['variable'],      _cell(align='left', bg=bg))
    apply(ws3, r, 2,  round(sr['adf_stat'], 4),  _cell(bg=bg), '0.0000')
    apply(ws3, r, 3,  round(sr['adf_p'], 4),     _cell(bg=bg), '0.0000')
    apply(ws3, r, 4,  sr['adf_lags'],             _cell(bg=bg))
    apply(ws3, r, 5,  sr['adf_conclusion'],
          _cell(bold=adf_ok, color=GREEN_DARK if adf_ok else RED_DARK, bg=bg))
    apply(ws3, r, 6,  round(sr['kpss_stat'], 4), _cell(bg=bg), '0.0000')
    apply(ws3, r, 7,  round(sr['kpss_p'], 4),    _cell(bg=bg), '0.0000')
    apply(ws3, r, 8,  sr['kpss_lags'],            _cell(bg=bg))
    apply(ws3, r, 9,  sr['kpss_conclusion'],
          _cell(bold=kpss_ok, color=GREEN_DARK if kpss_ok else RED_DARK, bg=bg))
    apply(ws3, r, 10, '✓ Confirmed' if both_ok else '⚠ Review',
          _cell(bold=True,
                color=GREEN_DARK if both_ok else RED_DARK,
                bg=GREEN_LIGHT if both_ok else RED_LIGHT))

autofit(ws3)
ws3.freeze_panes = 'A5'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 — DIAGNOSTICS (all tests)
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb_out.create_sheet('Diagnostics')
ws4.sheet_view.showGridLines = False

merge_hdr(ws4, 1, 1, 1, 6, 'Diagnostic Test Summary',
          _hdr(bg=BLUE_DARK, size=12))

diag_cols = ['Test', 'H₀', 'Statistic', 'p-value', 'Decision', 'Implication']
for c, h in enumerate(diag_cols, 1):
    apply(ws4, 2, c, h, _hdr(bg=GREY_HDR))

full_diag = [
    ('Breusch-Pagan', 'Homoskedasticity',
     f'LM = {bp_lm:.4f}', f'{bp_pval:.4f}',
     'Reject' if bp_pval < 0.05 else 'Fail',
     'Heteroskedasticity present → HAC needed'),
    ('Breusch-Godfrey (12L)', 'No serial correlation',
     f'LM = {bg_lm:.4f}', f'{bg_pval:.4f}',
     'Reject' if bg_pval < 0.05 else 'Fail',
     'Higher-order autocorrelation → HAC needed'),
    ('Durbin-Watson', 'No 1st-order autocorrel.',
     f'{dw:.4f}', '—',
     'No 1st-order AC' if 1.5 < dw < 2.5 else 'Inconclusive',
     'DW ≈ 2 consistent with BG detecting higher-order AC'),
    ('Jarque-Bera', 'Normal residuals',
     f'JB = {jb_stat:.2f}', f'{jb_pval:.4f}',
     'Reject' if jb_pval < 0.05 else 'Fail',
     f'Fat tails (kurt={kurt_val:.2f}); CLT valid at T={T}'),
    ('Ramsey RESET', 'Linear spec. adequate',
     f'F = {reset_f:.4f}', f'{reset_p:.4f}',
     'Reject' if reset_p < 0.05 else 'Fail',
     'Review nonlinear spec.' if reset_p < 0.05 else 'Linear form OK'),
    (f'Chow (break {break_date})', 'Stable coefficients',
     f'F = {chow_f:.4f}', f'{chow_p:.4f}',
     'Reject' if chow_p < 0.05 else 'Fail',
     'Structural break detected' if chow_p < 0.05
     else 'No break at chosen date'),
    ('Max VIF', 'VIF < 10 (no multicollinearity)',
     f'{max(vif_vals.values()):.2f}', '—',
     'OK ✓' if max(vif_vals.values()) < 10 else 'HIGH',
     'Refinery util. removed (VIF=13); remaining max VIF acceptable'),
]

for i, row_d in enumerate(full_diag):
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    r  = 3 + i
    for c, val in enumerate(row_d, 1):
        st = _cell(bg=bg, align='left' if c in (1, 2, 6) else 'center')
        if c == 5:  # Decision
            is_reject = 'Reject' in val
            st = _cell(bold=True, align='center',
                       color=RED_DARK if is_reject else GREEN_DARK, bg=bg)
        apply(ws4, r, c, val, st)

autofit(ws4)
ws4.column_dimensions['A'].width = 26
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['F'].width = 45
ws4.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 — BETA COEFFICIENTS
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb_out.create_sheet('Beta Coefficients')
ws5.sheet_view.showGridLines = False

merge_hdr(ws5, 1, 1, 1, 5,
          'Standardised (Beta) Coefficients — Relative Economic Importance',
          _hdr(bg=BLUE_DARK, size=12))
merge_hdr(ws5, 2, 1, 2, 5,
          'All variables standardised to zero mean, unit variance before regression',
          _hdr(bg=GREY_HDR, bold=False, size=9))

for c, h in enumerate(['Rank', 'Variable', 'Beta Coef', '|Beta|', 'HAC p-value'], 1):
    apply(ws5, 3, c, h, _hdr(bg=GREY_HDR))

beta_sorted = ols_std.params.drop('const').abs().sort_values(ascending=False)
for rank, (var, beta_abs) in enumerate(beta_sorted.items(), 1):
    r   = 3 + rank
    bg  = BLUE_LIGHT if rank % 2 == 0 else WHITE
    p   = final_model.pvalues[var]
    sig = p < 0.05
    apply(ws5, r, 1, rank, _cell(bg=bg))
    apply(ws5, r, 2, LABELS[var], _cell(align='left', bg=bg,
          bold=sig, color=RED_DARK if sig else '000000'))
    apply(ws5, r, 3, round(ols_std.params[var], 4),
          _cell(bg=GREEN_LIGHT if sig else bg,
                color=RED_DARK if ols_std.params[var] < 0 else GREEN_DARK,
                bold=sig), '+0.0000;-0.0000')
    apply(ws5, r, 4, round(beta_abs, 4), _cell(bg=bg), '0.0000')
    apply(ws5, r, 5, round(p, 4),
          _cell(bg=bg,
                color=RED_DARK if sig else '000000'), '0.0000')

autofit(ws5)
ws5.column_dimensions['B'].width = 32
ws5.freeze_panes = 'A4'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6 — DESCRIPTIVE STATISTICS
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb_out.create_sheet('Descriptive Statistics')
ws6.sheet_view.showGridLines = False

merge_hdr(ws6, 1, 1, 1, 8,
          f'Descriptive Statistics — All Series  |  T={T}  |  '
          f'{df.index[0]:%Y-%m} to {df.index[-1]:%Y-%m}',
          _hdr(bg=BLUE_DARK, size=12))

stat_hdrs = ['Variable', 'Mean', 'Std Dev', 'Min', 'P25', 'Median',
             'P75', 'Max', 'Skewness', 'Exc. Kurtosis']
for c, h in enumerate(stat_hdrs, 1):
    apply(ws6, 2, c, h, _hdr(bg=GREY_HDR))

FMT4 = '0.0000'
for i, col in enumerate(TARGET_COLS):
    s  = df[col].dropna()
    r  = 3 + i
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    vals = [LABELS[col],
            s.mean(), s.std(), s.min(),
            s.quantile(0.25), s.median(), s.quantile(0.75), s.max(),
            float(stats.skew(s)), float(stats.kurtosis(s))]
    for c, v in enumerate(vals, 1):
        fmt = FMT4 if isinstance(v, float) else None
        apply(ws6, r, c, round(v, 4) if isinstance(v, float) else v,
              _cell(align='left' if c == 1 else 'center', bg=bg), fmt)

autofit(ws6)
ws6.column_dimensions['A'].width = 32
ws6.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7 — RAW DATA
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb_out.create_sheet('Raw Data')
ws7.sheet_view.showGridLines = False

merge_hdr(ws7, 1, 1, 1, len(TARGET_COLS) + 1,
          'Cleaned Monthly Panel Data Used in Estimation',
          _hdr(bg=BLUE_DARK, size=12))

apply(ws7, 2, 1, 'Date', _hdr(bg=GREY_HDR))
for c, col in enumerate(TARGET_COLS, 2):
    apply(ws7, 2, c, LABELS[col], _hdr(bg=GREY_HDR, size=8))

df_out = df[TARGET_COLS].reset_index()
for i, row_data in enumerate(df_out.itertuples(index=False), 3):
    bg = BLUE_LIGHT if i % 2 == 0 else WHITE
    for c, val in enumerate(row_data, 1):
        fmt = '0.0000' if isinstance(val, float) else (
              'YYYY-MM' if isinstance(val, (pd.Timestamp, datetime.datetime))
              else None)
        apply(ws7, i, c,
              val.date() if isinstance(val, (pd.Timestamp,)) else val,
              _cell(bg=bg, align='center'), fmt)

autofit(ws7, min_w=10, max_w=20)
ws7.column_dimensions['A'].width = 12
ws7.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
wb_out.save(OUTPUT_XLSX)
print(f'\nSaved: {OUTPUT_XLSX}')
print(f'  Sheets: {wb_out.sheetnames}')


# ═══════════════════════════════════════════════════════════════════════════════
# 11.  CONSOLE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

print('\n' + '='*72)
print('FINAL SUMMARY')
print('='*72)
sig_vars   = [v for v in REGRESSORS if final_model.pvalues[v] < 0.05]
insig_vars = [v for v in REGRESSORS if final_model.pvalues[v] >= 0.05]
print(f'  R²={ols.rsquared:.4f}  Adj-R²={ols.rsquared_adj:.4f}  '
      f'F={ols.fvalue:.2f} (p={ols.f_pvalue:.4f})')
print(f'  Heteroskedast. : BP p={bp_pval:.4f}  '
      f'→ {"HAC applied" if bp_pval < 0.05 else "Not detected"}')
print(f'  Autocorrelation: BG p={bg_pval:.4f}  '
      f'→ {"HAC applied" if bg_pval < 0.05 else "Not detected"}')
print(f'  Normality      : JB p={jb_pval:.4f}  '
      f'→ {"Non-normal (fat tails)" if jb_pval < 0.05 else "Approx. normal"}')
print(f'\n  Significant regressors (HAC p<0.05):')
for v in sig_vars:
    c_ = final_model.params[v]
    p_ = final_model.pvalues[v]
    print(f'    {LABELS[v]:<32}  coef={c_:+.4f}  p={p_:.4f}  '
          f'({"pos" if c_ > 0 else "neg"})')
print(f'\n  Insignificant (p≥0.05):')
for v in insig_vars:
    print(f'    {LABELS[v]:<32}  p={final_model.pvalues[v]:.4f}')

print('\nDone.')
print('Output files:')
print('  residual_diagnostics.png')
print('  coefficient_plot.png')
print(f'  {OUTPUT_XLSX}  ({wb_out.sheetnames})')